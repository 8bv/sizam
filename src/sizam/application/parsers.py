import logging
from typing import Optional

from openpyxl import load_workbook

from ..application.schemas import UntiSetStatusDTO
from ..services.consts import STATUSES_MAPPING, UntiCourseStatus


logger = logging.getLogger(__name__)


class InvalidHeader(Exception):
    'Неверный формат первой строки в файле'


def get_data(
        path: str,
        sheet: Optional[str] = None,
        status: UntiCourseStatus = UntiCourseStatus.UNKNOWN,
) -> list[UntiSetStatusDTO]:
    workbook = load_workbook(path, read_only=True)
    if sheet is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet]

    it = worksheet.iter_rows(values_only=True)

    header = next(it)
    if not any(header):
        raise InvalidHeader("Первая строка не найдена")

    header_template = ["unti_id", "course_id"]
    if status is UntiCourseStatus.UNKNOWN:
        header_template.append("status")

    if [v.lower() for v in header] != header_template:
        raise InvalidHeader("Формат первой строки не соответствует шаблону: %s" % ", ".join(header_template))

    result = []
    for row in it:
        logger.debug("%s", row)
        if status is UntiCourseStatus.UNKNOWN:
            status = STATUSES_MAPPING[row[2].lower()]

        result.append(
            UntiSetStatusDTO(
                unti_id=int(row[0]),
                course_id=int(row[1]),
                status=status,
            )
        )
    return result
